import os
import shutil
from pathlib import Path

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def _expand_path(path):
    return Path(path).expanduser()


def read_file(path):
    try:
        return _expand_path(path).read_text(encoding="utf-8")
    except Exception as e:
        return str(e)


def write_file(path, content):
    try:
        target_path = _expand_path(path)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target_path.write_text(content, encoding="utf-8")
        return f"File written: {target_path}"
    except Exception as e:
        return str(e)


def append_file(path, content):
    try:
        target_path = _expand_path(path)
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with target_path.open("a", encoding="utf-8") as file_handle:
            file_handle.write(content)
        return f"File updated: {target_path}"
    except Exception as e:
        return str(e)


def delete_file(path):
    try:
        target_path = _expand_path(path)
        target_path.unlink()
        return f"Deleted file: {target_path}"
    except Exception as e:
        return str(e)


def copy_file(src, dst):
    try:
        source_path = _expand_path(src)
        destination_path = _expand_path(dst)
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination_path)
        return f"Copied file to: {destination_path}"
    except Exception as e:
        return str(e)


def move_file(src, dst):
    try:
        source_path = _expand_path(src)
        destination_path = _expand_path(dst)
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source_path), str(destination_path))
        return f"Moved file to: {destination_path}"
    except Exception as e:
        return str(e)


def rename_file(old, new):
    return move_file(old, new)


def get_current_directory():
    try:
        return os.getcwd()
    except Exception as e:
        return str(e)


def list_files(path="."):
    try:
        target_path = _expand_path(path)
        files = sorted(item.name for item in target_path.iterdir())
        return "\n".join(files) if files else "No files found."
    except Exception as e:
        return str(e)


def search_file(folder, name):
    try:
        target_path = _expand_path(folder)
        matches = [str(path) for path in target_path.rglob("*") if name.lower() in path.name.lower()]
        return "\n".join(matches) if matches else f"No files found matching '{name}'."
    except Exception as e:
        return str(e)


def get_file_info(path):
    try:
        target_path = _expand_path(path)
        stat = target_path.stat()
        item_type = "Folder" if target_path.is_dir() else "File"
        return (
            f"Path: {target_path}\n"
            f"Type: {item_type}\n"
            f"Size: {stat.st_size} bytes\n"
            f"Modified: {stat.st_mtime}"
        )
    except Exception as e:
        return str(e)


def read_pdf(path):
    if PdfReader is None:
        return "PyPDF2 is not installed"

    try:
        reader = PdfReader(str(_expand_path(path)))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages).strip()
        return text or "No text found in PDF."
    except Exception as e:
        return str(e)


def read_word(path):
    if Document is None:
        return "python-docx is not installed"

    try:
        document = Document(str(_expand_path(path)))
        paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
        return "\n".join(paragraphs) if paragraphs else "No text found in Word document."
    except Exception as e:
        return str(e)


def read_excel(path):
    if load_workbook is None:
        return "openpyxl is not installed"

    try:
        workbook = load_workbook(filename=str(_expand_path(path)), data_only=True)
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            rows.append(" | ".join(values).rstrip())
        return "\n".join(rows) if rows else "Excel file is empty."
    except Exception as e:
        return str(e)


def zip_folder(folder, output):
    try:
        source_path = _expand_path(folder)
        output_path = _expand_path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        archive_base = output_path.with_suffix("")
        archive_path = shutil.make_archive(str(archive_base), "zip", root_dir=str(source_path))
        return f"Created ZIP: {archive_path}"
    except Exception as e:
        return str(e)
